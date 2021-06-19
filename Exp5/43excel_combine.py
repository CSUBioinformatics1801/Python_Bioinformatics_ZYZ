import openpyxl
import os
def excel_srt():
    wb1=openpyxl.load_workbook('D:\\Remain\\7.python及其生物信息学应用\\实践\\实践5\\sheet1.xlsx')
    wb2=openpyxl.load_workbook('D:\\Remain\\7.python及其生物信息学应用\\实践\\实践5\\sheet2.xlsx')
    sheet1=wb1.active
    row1=sheet1.max_row
    sheet2=wb2.active
    for i in range (2,51):
        id2=sheet2.cell(i,1).value
        for j in range(2,row1+1):
            id1=sheet1.cell(j,1).value
            if(id1==id2):
                for k in range(5):
                    sheet1.cell(j,4+k).value=sheet2.cell(i,2+k).value
    wb1.save('D:\\Remain\\7.python及其生物信息学应用\\实践\\实践5\\sum.xlsx')
if __name__ == "__main__":
    os.path.abspath(__file__)
    excel_srt()